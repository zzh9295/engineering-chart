"""导出器：CSV、图片、Excel、JSON"""
import csv
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from gantt_app.config import (
    MAX_IMAGE_WIDTH, MAX_IMAGE_HEIGHT, MAX_EXCEL_COLUMNS,
    TASK_COLORS, EXCEL_FILLS, FONT_CANDIDATES, THEME
)
from gantt_app.models import parse_tasks_for_export
from gantt_app.utils import compute_layout, prepare_task_data


class ExportManager:
    """统一导出管理器，负责所有文件导出功能"""

    def __init__(self, app):
        self.app = app

    def _get_font(self, size):
        """按候选列表获取可用字体"""
        for name in FONT_CANDIDATES:
            try:
                return ImageFont.truetype(name, size)
            except Exception:
                continue
        return ImageFont.load_default()

    def _show_progress(self, title, message):
        """显示导出进度窗口"""
        progress = tk.Toplevel(self.app.root)
        progress.title(title)
        progress.geometry("300x80")
        progress.transient(self.app.root)
        progress.grab_set()
        ttk.Label(progress, text=message).pack(pady=10)
        pb = ttk.Progressbar(progress, mode='indeterminate')
        pb.pack(fill='x', padx=20)
        pb.start()
        return progress

    def _safe_destroy(self, widget):
        """安全销毁控件"""
        try:
            widget.destroy()
        except Exception:
            pass

    def export_csv(self):
        """导出 CSV 文件"""
        tasks = parse_tasks_for_export(self.app.task_rows)
        if not tasks:
            messagebox.showwarning("提示", "没有有效的任务数据可导出")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")],
            title="导出CSV文件"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['项目名称', self.app.project_name.get()])
                writer.writerow(['项目编号', self.app.project_code.get()])
                writer.writerow([])
                writer.writerow(['序号', '任务名称', '开始日期', '结束日期', '工期(天)', '任务类型'])
                for task in tasks:
                    writer.writerow([
                        task['序号'], task['任务名称'], task['开始日期'],
                        task['结束日期'], task['工期(天)'], task['任务类型']
                    ])
            messagebox.showinfo("成功", f"CSV文件已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")

    def export_image(self):
        """导出图片文件"""
        tasks = parse_tasks_for_export(self.app.task_rows)
        if not tasks:
            messagebox.showwarning("提示", "没有有效的任务数据可导出")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG图片", "*.png"), ("JPG图片", "*.jpg"), ("所有文件", "*.*")],
            title="导出图片"
        )

        if not file_path:
            return

        progress = self._show_progress("导出进度", "正在生成横道图图片，请稍候...")
        self.app.root.after(200, lambda: self._do_export_image(file_path, tasks, progress))

    def _do_export_image(self, file_path, tasks, progress):
        try:
            # 统一转换为 compute_layout 所需的 start/end 格式
            normalized = [
                {
                    'name': t['任务名称'],
                    'start': t['start_dt'],
                    'end': t['end_dt'],
                    'type': t['任务类型']
                }
                for t in tasks
            ]

            layout = compute_layout(normalized, zoom_level=1.0)
            prepared = prepare_task_data(normalized, layout.min_date)

            image = Image.new('RGB', (layout.width, layout.height), THEME['bg_card'])
            draw = ImageDraw.Draw(image)

            font_title = self._get_font(14)
            font_header = self._get_font(10)
            font_task = self._get_font(max(8, min(11, layout.row_height // 3)))
            font_day = self._get_font(max(6, min(9, layout.day_width // 3)))

            padding = layout.padding
            draw.text((padding, padding), f"项目名称: {self.app.project_name.get()}", font=font_title, fill=THEME['text_primary'])
            draw.text((padding, padding + 20), f"项目编号: {self.app.project_code.get()}", font=font_title, fill=THEME['text_primary'])

            header_y = padding + 45
            draw.rectangle(
                [padding, header_y, padding + layout.left_width, header_y + layout.header_height],
                fill=THEME['header_bg'], outline=THEME['border']
            )
            draw.text(
                (padding + layout.left_width / 2, header_y + layout.header_height / 2),
                "任务名称", font=font_header, fill=THEME['text_primary'], anchor='mm'
            )

            current = layout.min_date
            x = padding + layout.left_width
            while current <= layout.max_date:
                draw.rectangle(
                    [x, header_y, x + layout.day_width, header_y + layout.header_height],
                    fill='#f9fafb', outline=THEME['border']
                )
                if layout.day_width >= 10:
                    draw.text(
                        (x + layout.day_width / 2, header_y + layout.header_height / 2),
                        current.strftime("%m/%d"), font=font_day, fill=THEME['text_secondary'], anchor='mm'
                    )
                elif layout.day_width >= 6:
                    draw.text(
                        (x + layout.day_width / 2, header_y + layout.header_height / 2),
                        current.strftime("%m-%d"), font=font_day, fill=THEME['text_secondary'], anchor='mm'
                    )
                current += timedelta(days=1)
                x += layout.day_width

            for i, task in enumerate(prepared):
                y = header_y + layout.header_height + i * layout.row_height

                draw.rectangle(
                    [padding, y, padding + layout.left_width, y + layout.row_height],
                    fill=THEME['bg_card'], outline=THEME['border']
                )
                draw.text(
                    (padding + 10, y + layout.row_height / 2),
                    task['name'], font=font_task, fill=THEME['text_primary'], anchor='lm'
                )

                bar_x = padding + layout.left_width + task['start_offset'] * layout.day_width
                bar_width = task['duration'] * layout.day_width
                bar_width = max(1, bar_width)

                color = TASK_COLORS.get(task['type'], TASK_COLORS['normal'])
                bar_top = y + layout.bar_margin
                bar_bottom = y + layout.row_height - layout.bar_margin
                draw.rectangle([bar_x, bar_top, bar_x + bar_width, bar_bottom], fill=color)

                if bar_width >= max(24, layout.day_width):
                    draw.text(
                        (bar_x + bar_width / 2, y + layout.row_height / 2),
                        f"{task['duration']}天", font=font_day, fill='white', anchor='mm'
                    )

            image.save(file_path)
            self._safe_destroy(progress)
            messagebox.showinfo("成功", f"图片已导出到:\n{file_path}")
        except Exception as e:
            self._safe_destroy(progress)
            messagebox.showerror("错误", f"导出失败: {str(e)}")

    def export_excel(self):
        """导出 Excel 文件"""
        tasks = parse_tasks_for_export(self.app.task_rows)
        if not tasks:
            messagebox.showwarning("提示", "没有有效的任务数据可导出")
            return

        try:
            min_date = min(t['start_dt'] for t in tasks)
            max_date = max(t['end_dt'] for t in tasks)
            total_days = (max_date - min_date).days + 1

            if total_days > MAX_EXCEL_COLUMNS - 3:
                messagebox.showwarning("提示", f"时间跨度太大（{total_days}天），超出Excel列数限制")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "横道图"

            header_font = Font(bold=True, size=11)
            title_font = Font(bold=True, size=14)
            day_font = Font(size=9)
            task_font = Font(size=10)

            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            colors = {
                k: PatternFill(start_color=v, end_color=v, fill_type='solid')
                for k, v in EXCEL_FILLS.items()
            }
            header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            day_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

            ws.merge_cells('A1:C1')
            ws['A1'] = f"项目名称: {self.app.project_name.get()}"
            ws['A1'].font = title_font

            ws.merge_cells('A2:C2')
            ws['A2'] = f"项目编号: {self.app.project_code.get()}"
            ws['A2'].font = title_font

            row_offset = 4
            ws['A' + str(row_offset)] = "任务名称"
            cell = ws['A' + str(row_offset)]
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            current = min_date
            col = 2
            while current <= max_date:
                cell = ws.cell(row=row_offset, column=col, value=current.strftime("%m/%d"))
                cell.font = day_font
                cell.fill = day_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                current += timedelta(days=1)
                col += 1

            for i, task in enumerate(tasks):
                row = row_offset + 1 + i

                cell = ws.cell(row=row, column=1, value=task['任务名称'])
                cell.font = task_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                start_offset = (task['start_dt'] - min_date).days
                duration = (task['end_dt'] - task['start_dt']).days + 1

                for j in range(total_days):
                    cell = ws.cell(row=row, column=2 + j)
                    cell.border = thin_border
                    if start_offset <= j < start_offset + duration:
                        cell.fill = colors.get(task['任务类型'], colors['normal'])

                info_cell = ws.cell(row=row, column=2 + total_days, value=f"{duration}天")
                info_cell.font = task_font
                info_cell.border = thin_border
                info_cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.column_dimensions['A'].width = 25
            for i in range(2, col + 1):
                ws.column_dimensions[get_column_letter(i)].width = 6

            ws.row_dimensions[row_offset].height = 30
            for i in range(row_offset + 1, row_offset + 1 + len(tasks)):
                ws.row_dimensions[i].height = 25

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                title="导出Excel文件"
            )

            if not file_path:
                return

            try:
                wb.save(file_path)
                messagebox.showinfo("成功", f"Excel文件已导出到:\n{file_path}")
            except PermissionError:
                messagebox.showerror("错误", "无法保存文件，可能是文件已被其他程序打开")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")

        except Exception as e:
            messagebox.showerror("错误", f"生成Excel失败: {str(e)}")

    def export_json(self):
        """导出 JSON 文件"""
        tasks = parse_tasks_for_export(self.app.task_rows)
        if not tasks:
            messagebox.showwarning("提示", "没有有效的任务数据可导出")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")],
            title="导出JSON文件"
        )

        if not file_path:
            return

        progress = self._show_progress("导出进度", "正在导出JSON数据，请稍候...")
        self.app.root.after(200, lambda: self._do_export_json(file_path, tasks, progress))

    def _do_export_json(self, file_path, tasks, progress):
        try:
            data = {
                '项目名称': self.app.project_name.get(),
                '项目编号': self.app.project_code.get(),
                '导出时间': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                '任务总数': len(tasks),
                '任务列表': tasks
            }

            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self._safe_destroy(progress)
            messagebox.showinfo("成功", f"JSON文件已导出到:\n{file_path}")
        except Exception as e:
            self._safe_destroy(progress)
            messagebox.showerror("错误", f"导出失败: {str(e)}")
